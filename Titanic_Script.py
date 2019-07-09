# This script takes the data from the 'Titanic' data set
# and writes it to two separate files, one where the column
# order is reversed and one where every other column is removed.
# For some reason the Titanic data set has a file name of 'train.xlsx'.
import openpyxl
from openpyxl import Workbook
from string import ascii_uppercase


def titanic_script():

    wb1 = openpyxl.load_workbook("train.xlsx")
    ws1 = wb1.active

    #  create excel files
    wb2 = Workbook("Reversed_Columns.xlsx")
    wb2.save("Reversed_Columns.xlsx")
    wb2 = openpyxl.load_workbook("Reversed_Columns.xlsx")
    ws2 = wb2.active

    wb3 = Workbook("Removed_Every_Other_Column.xlsx")
    wb3.save("Removed_Every_Other_Column.xlsx")
    wb3 = openpyxl.load_workbook("Removed_Every_Other_Column.xlsx")
    ws3 = wb3.active

    total_cols = len(tuple(ws1.columns))

    # first 'for' loop iterates for each column, the second iterates through each
    # cell in that column and writes that value to the new file in it's new position
    for col_num in range(total_cols):
        for i in range(len(ws1[ascii_uppercase[col_num]])):
            ws2[ascii_uppercase[total_cols-col_num-1] + str(i+1)] = ws1[ascii_uppercase[col_num] + str(i+1)].value


    # writes data to new file only for every other column
    for col_num in range(total_cols):
        if col_num % 2 == 0:
            for i in range(len(ws1[ascii_uppercase[col_num]])):
                ws3[ascii_uppercase[col_num] + str(i + 1)].value = ws1[ascii_uppercase[col_num] + str(i+1)].value

    wb2.save("Reversed_Columns.xlsx")
    wb3.save("Removed_Every_Other_Column.xlsx")

